import streamlit as st
import pandas as pd
import numpy as np
import io
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Set page configuration
st.set_page_config(
    page_title="Advanced Reservoir Reserves Calculator",
    page_icon="🛢️",
    layout="wide"
)

# Helper function for safe Excel sheet names
def get_unique_sheet_name(base_name, used_names):
    """Create a safe, unique Excel sheet name"""
    # Replace invalid characters
    for char in [':', '\\', '/', '?', '*', '[', ']']:
        base_name = base_name.replace(char, '_')
    
    # Truncate to 31 characters (Excel limit)
    name = base_name[:31]
    
    # Ensure uniqueness
    if name not in used_names:
        used_names.append(name)
        return name
    
    # Add number suffix if name exists
    i = 1
    while f"{name}_{i}" in used_names:
        i += 1
    unique_name = f"{name}_{i}"
    used_names.append(unique_name)
    return unique_name

# Title and description
st.title("🛢️ Advanced Reservoir Reserves Calculator")
st.markdown("Calculate Original Oil In Place (OOIP) with comprehensive sensitivity analysis")

# Input section
st.header("Input Parameters")
st.markdown("Enter your reservoir properties below:")

# Create columns for better layout
col1, col2 = st.columns(2)

with col1:
    # Area input
    area = st.number_input(
        "Reservoir Area (acres)",
        min_value=10,
        max_value=100000,
        value=1000,
        step=10,
        help="Total area of the reservoir in acres"
    )
    
    # Thickness input
    thickness = st.number_input(
        "Net Pay Thickness (ft)",
        min_value=5,
        max_value=1000,
        value=50,
        step=1,
        help="Average net pay thickness in feet"
    )
    
    # Porosity input
    porosity = st.slider(
        "Porosity (%)",
        min_value=5.0,
        max_value=40.0,
        value=20.0,
        step=0.5,
        help="Average porosity as a percentage"
    )

with col2:
    # Water saturation input
    water_saturation = st.slider(
        "Water Saturation (%)",
        min_value=5.0,
        max_value=50.0,
        value=25.0,
        step=0.5,
        help="Average water saturation as a percentage"
    )
    
    # Oil FVF input
    oil_fvf = st.number_input(
        "Oil Formation Volume Factor (rb/stb)",
        min_value=1.0,
        max_value=2.5,
        value=1.3,
        step=0.01,
        help="Oil formation volume factor in reservoir barrels per stock tank barrel"
    )
    
    # Recovery factor input
    recovery_factor = st.slider(
        "Recovery Factor (%)",
        min_value=5.0,
        max_value=60.0,
        value=30.0,
        step=1.0,
        help="Estimated recovery factor as a percentage"
    )

# Calculate reserves
st.header("Results")

# Convert percentages to decimals
porosity_decimal = porosity / 100
water_saturation_decimal = water_saturation / 100
recovery_factor_decimal = recovery_factor / 100

# Calculate OOIP using volumetric method
# Formula: OOIP = (7758 * A * h * φ * (1-Sw)) / Bo
ooip = (7758 * area * thickness * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf

# Calculate Recoverable Reserves
recoverable_reserves = ooip * recovery_factor_decimal

# Display results
col1, col2, col3 = st.columns(3)

with col1:
    st.metric(
        "Original Oil In Place (OOIP)",
        f"{ooip:,.0f} bbl",
        help="Total oil initially in place"
    )

with col2:
    st.metric(
        "Recoverable Reserves",
        f"{recoverable_reserves:,.0f} bbl",
        help="Estimated recoverable oil volume"
    )

with col3:
    st.metric(
        "Recovery Factor",
        f"{recovery_factor}%",
        help="Percentage of OOIP that can be recovered"
    )

# Sensitivity analysis section
st.header("Comprehensive Sensitivity Analysis")
st.markdown("Analyze how changes in each parameter affect OOIP and Recoverable Reserves")

# Create tabs for different sensitivity views
tab1, tab2, tab3 = st.tabs(["Single Parameter Analysis", "Tornado Plot", "Multi-Parameter Matrix"])

# Tab 1: Single Parameter Analysis
with tab1:
    st.subheader("Single Parameter Sensitivity")
    
    # Select parameter to analyze
    param_to_analyze = st.selectbox(
        "Select parameter to analyze",
        ["Area", "Thickness", "Porosity", "Water Saturation", "Oil FVF", "Recovery Factor"]
    )
    
    # Define parameter ranges
    param_ranges = {
        "Area": np.linspace(area * 0.5, area * 1.5, 100),
        "Thickness": np.linspace(thickness * 0.5, thickness * 1.5, 100),
        "Porosity": np.linspace(porosity * 0.5, porosity * 1.5, 100),
        "Water Saturation": np.linspace(water_saturation * 0.5, water_saturation * 1.5, 100),
        "Oil FVF": np.linspace(oil_fvf * 0.8, oil_fvf * 1.2, 100),
        "Recovery Factor": np.linspace(recovery_factor * 0.5, recovery_factor * 1.5, 100)
    }
    
    # Calculate OOIP and Recoverable Reserves for each parameter value
    ooip_values = []
    recoverable_values = []
    
    for value in param_ranges[param_to_analyze]:
        if param_to_analyze == "Area":
            temp_ooip = (7758 * value * thickness * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
        elif param_to_analyze == "Thickness":
            temp_ooip = (7758 * area * value * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
        elif param_to_analyze == "Porosity":
            temp_ooip = (7758 * area * thickness * (value/100) * (1 - water_saturation_decimal)) / oil_fvf
        elif param_to_analyze == "Water Saturation":
            temp_ooip = (7758 * area * thickness * porosity_decimal * (1 - (value/100))) / oil_fvf
        elif param_to_analyze == "Oil FVF":
            temp_ooip = (7758 * area * thickness * porosity_decimal * (1 - water_saturation_decimal)) / value
        elif param_to_analyze == "Recovery Factor":
            temp_ooip = ooip  # OOIP doesn't change with recovery factor
            
        if param_to_analyze == "Recovery Factor":
            temp_recoverable = ooip * (value/100)
        else:
            temp_recoverable = temp_ooip * recovery_factor_decimal
            
        ooip_values.append(temp_ooip)
        recoverable_values.append(temp_recoverable)
    
    # Create DataFrame for plotting
    sensitivity_df = pd.DataFrame({
        f'{param_to_analyze}': param_ranges[param_to_analyze],
        'OOIP (bbl)': ooip_values,
        'Recoverable Reserves (bbl)': recoverable_values
    })
    
    # Plot the sensitivity analysis
    fig = px.line(
        sensitivity_df, 
        x=f'{param_to_analyze}', 
        y=['OOIP (bbl)', 'Recoverable Reserves (bbl)'],
        title=f'Sensitivity Analysis: {param_to_analyze}',
        labels={'value': 'Reserves (bbl)', f'{param_to_analyze}': param_to_analyze}
    )
    
    # Add vertical line for base case
    fig.add_vline(
        x=eval(param_to_analyze.lower().replace(' ', '_')), 
        line_dash="dash", 
        line_color="red",
        annotation_text="Base Case"
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Calculate and display sensitivity metrics
    base_ooip = ooip
    base_recoverable = recoverable_reserves
    
    min_ooip = min(ooip_values)
    max_ooip = max(ooip_values)
    min_recoverable = min(recoverable_values)
    max_recoverable = max(recoverable_values)
    
    col1, col2 = st.columns(2)
    with col1:
        st.metric(
            f"OOIP Range ({param_to_analyze})",
            f"{min_ooip:,.0f} - {max_ooip:,.0f} bbl",
            delta=f"{((max_ooip - min_ooip) / base_ooip * 100):.1f}% variation"
        )
    with col2:
        st.metric(
            f"Recoverable Range ({param_to_analyze})",
            f"{min_recoverable:,.0f} - {max_recoverable:,.0f} bbl",
            delta=f"{((max_recoverable - min_recoverable) / base_recoverable * 100):.1f}% variation"
        )

# Tab 2: Tornado Plot
with tab2:
    st.subheader("Tornado Plot - Parameter Impact Comparison")
    
    # Define variation percentage (±20%)
    variation_percent = 0.2
    
    # Calculate base case
    base_ooip = ooip
    base_recoverable = recoverable_reserves
    
    # Calculate low and high values for each parameter
    param_low_values = {
        "Area": area * (1 - variation_percent),
        "Thickness": thickness * (1 - variation_percent),
        "Porosity": porosity * (1 - variation_percent),
        "Water Saturation": water_saturation * (1 - variation_percent),
        "Oil FVF": oil_fvf * (1 - variation_percent),
        "Recovery Factor": recovery_factor * (1 - variation_percent)
    }
    
    param_high_values = {
        "Area": area * (1 + variation_percent),
        "Thickness": thickness * (1 + variation_percent),
        "Porosity": porosity * (1 + variation_percent),
        "Water Saturation": water_saturation * (1 + variation_percent),
        "Oil FVF": oil_fvf * (1 + variation_percent),
        "Recovery Factor": recovery_factor * (1 + variation_percent)
    }
    
    # Calculate OOIP and Recoverable Reserves for low and high values
    ooip_low = {}
    ooip_high = {}
    recoverable_low = {}
    recoverable_high = {}
    
    for param in param_low_values:
        if param == "Area":
            ooip_low[param] = (7758 * param_low_values[param] * thickness * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
            ooip_high[param] = (7758 * param_high_values[param] * thickness * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
        elif param == "Thickness":
            ooip_low[param] = (7758 * area * param_low_values[param] * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
            ooip_high[param] = (7758 * area * param_high_values[param] * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
        elif param == "Porosity":
            ooip_low[param] = (7758 * area * thickness * (param_low_values[param]/100) * (1 - water_saturation_decimal)) / oil_fvf
            ooip_high[param] = (7758 * area * thickness * (param_high_values[param]/100) * (1 - water_saturation_decimal)) / oil_fvf
        elif param == "Water Saturation":
            ooip_low[param] = (7758 * area * thickness * porosity_decimal * (1 - (param_low_values[param]/100))) / oil_fvf
            ooip_high[param] = (7758 * area * thickness * porosity_decimal * (1 - (param_high_values[param]/100))) / oil_fvf
        elif param == "Oil FVF":
            ooip_low[param] = (7758 * area * thickness * porosity_decimal * (1 - water_saturation_decimal)) / param_high_values[param]  # Inverse relationship
            ooip_high[param] = (7758 * area * thickness * porosity_decimal * (1 - water_saturation_decimal)) / param_low_values[param]   # Inverse relationship
        elif param == "Recovery Factor":
            ooip_low[param] = base_ooip  # OOIP doesn't change with recovery factor
            ooip_high[param] = base_ooip
            
        if param == "Recovery Factor":
            recoverable_low[param] = base_ooip * (param_low_values[param]/100)
            recoverable_high[param] = base_ooip * (param_high_values[param]/100)
        else:
            recoverable_low[param] = ooip_low[param] * recovery_factor_decimal
            recoverable_high[param] = ooip_high[param] * recovery_factor_decimal
    
    # Create tornado plot for OOIP
    fig_ooip = go.Figure()
    
    for param in param_low_values:
        fig_ooip.add_trace(go.Scatter(
            x=[ooip_low[param], ooip_high[param]],
            y=[param, param],
            mode='lines+markers',
            name=param,
            line=dict(width=10),
            marker=dict(size=10)
        ))
    
    # Add vertical line for base case
    fig_ooip.add_vline(
        x=base_ooip, 
        line_dash="dash", 
        line_color="red",
        annotation_text="Base Case"
    )
    
    fig_ooip.update_layout(
        title="Tornado Plot: Impact on OOIP",
        xaxis_title="OOIP (bbl)",
        yaxis_title="Parameters",
        height=500,
        showlegend=False
    )
    
    st.plotly_chart(fig_ooip, use_container_width=True)
    
    # Create tornado plot for Recoverable Reserves
    fig_recoverable = go.Figure()
    
    for param in param_low_values:
        fig_recoverable.add_trace(go.Scatter(
            x=[recoverable_low[param], recoverable_high[param]],
            y=[param, param],
            mode='lines+markers',
            name=param,
            line=dict(width=10),
            marker=dict(size=10)
        ))
    
    # Add vertical line for base case
    fig_recoverable.add_vline(
        x=base_recoverable, 
        line_dash="dash", 
        line_color="red",
        annotation_text="Base Case"
    )
    
    fig_recoverable.update_layout(
        title="Tornado Plot: Impact on Recoverable Reserves",
        xaxis_title="Recoverable Reserves (bbl)",
        yaxis_title="Parameters",
        height=500,
        showlegend=False
    )
    
    st.plotly_chart(fig_recoverable, use_container_width=True)

# Tab 3: Multi-Parameter Matrix
with tab3:
    st.subheader("Multi-Parameter Sensitivity Matrix")
    
    # Select two parameters to analyze
    col1, col2 = st.columns(2)
    with col1:
        param1 = st.selectbox(
            "Select first parameter",
            ["Area", "Thickness", "Porosity", "Water Saturation", "Oil FVF", "Recovery Factor"],
            key="param1"
        )
    with col2:
        param2 = st.selectbox(
            "Select second parameter",
            ["Area", "Thickness", "Porosity", "Water Saturation", "Oil FVF", "Recovery Factor"],
            key="param2"
        )
    
    # Define ranges for the two parameters
    param1_range = np.linspace(
        eval(param1.lower().replace(' ', '_')) * 0.8,
        eval(param1.lower().replace(' ', '_')) * 1.2,
        10
    )
    
    param2_range = np.linspace(
        eval(param2.lower().replace(' ', '_')) * 0.8,
        eval(param2.lower().replace(' ', '_')) * 1.2,
        10
    )
    
    # Create a matrix of OOIP values
    ooip_matrix = np.zeros((len(param1_range), len(param2_range)))
    recoverable_matrix = np.zeros((len(param1_range), len(param2_range)))
    
    for i, p1 in enumerate(param1_range):
        for j, p2 in enumerate(param2_range):
            # Calculate OOIP with modified parameters
            if param1 == "Area" and param2 == "Thickness":
                temp_ooip = (7758 * p1 * p2 * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
            elif param1 == "Area" and param2 == "Porosity":
                temp_ooip = (7758 * p1 * thickness * (p2/100) * (1 - water_saturation_decimal)) / oil_fvf
            elif param1 == "Area" and param2 == "Water Saturation":
                temp_ooip = (7758 * p1 * thickness * porosity_decimal * (1 - (p2/100))) / oil_fvf
            elif param1 == "Area" and param2 == "Oil FVF":
                temp_ooip = (7758 * p1 * thickness * porosity_decimal * (1 - water_saturation_decimal)) / p2
            elif param1 == "Area" and param2 == "Recovery Factor":
                temp_ooip = (7758 * p1 * thickness * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
            elif param1 == "Thickness" and param2 == "Porosity":
                temp_ooip = (7758 * area * p1 * (p2/100) * (1 - water_saturation_decimal)) / oil_fvf
            elif param1 == "Thickness" and param2 == "Water Saturation":
                temp_ooip = (7758 * area * p1 * porosity_decimal * (1 - (p2/100))) / oil_fvf
            elif param1 == "Thickness" and param2 == "Oil FVF":
                temp_ooip = (7758 * area * p1 * porosity_decimal * (1 - water_saturation_decimal)) / p2
            elif param1 == "Thickness" and param2 == "Recovery Factor":
                temp_ooip = (7758 * area * p1 * porosity_decimal * (1 - water_saturation_decimal)) / oil_fvf
            elif param1 == "Porosity" and param2 == "Water Saturation":
                temp_ooip = (7758 * area * thickness * (p1/100) * (1 - (p2/100))) / oil_fvf
            elif param1 == "Porosity" and param2 == "Oil FVF":
                temp_ooip = (7758 * area * thickness * (p1/100) * (1 - water_saturation_decimal)) / p2
            elif param1 == "Porosity" and param2 == "Recovery Factor":
                temp_ooip = (7758 * area * thickness * (p1/100) * (1 - water_saturation_decimal)) / oil_fvf
            elif param1 == "Water Saturation" and param2 == "Oil FVF":
                temp_ooip = (7758 * area * thickness * porosity_decimal * (1 - (p1/100))) / p2
            elif param1 == "Water Saturation" and param2 == "Recovery Factor":
                temp_ooip = (7758 * area * thickness * porosity_decimal * (1 - (p1/100))) / oil_fvf
            elif param1 == "Oil FVF" and param2 == "Recovery Factor":
                temp_ooip = (7758 * area * thickness * porosity_decimal * (1 - water_saturation_decimal)) / p1
            else:
                # Handle same parameter selected twice or other combinations
                temp_ooip = base_ooip
            
            # Calculate recoverable reserves
            if param1 == "Recovery Factor" and param2 == "Recovery Factor":
                temp_recoverable = temp_ooip * (p1/100)  # Using param1 value
            elif param1 == "Recovery Factor":
                temp_recoverable = temp_ooip * (p1/100)
            elif param2 == "Recovery Factor":
                temp_recoverable = temp_ooip * (p2/100)
            else:
                temp_recoverable = temp_ooip * recovery_factor_decimal
            
            ooip_matrix[i, j] = temp_ooip
            recoverable_matrix[i, j] = temp_recoverable
    
    # Create heatmaps
    fig_ooip_heatmap = px.imshow(
        ooip_matrix,
        x=[f"{p:.1f}" for p in param2_range],
        y=[f"{p:.1f}" for p in param1_range],
        labels=dict(x=f"{param2}", y=f"{param1}", color="OOIP (bbl)"),
        color_continuous_scale="Viridis",
        title=f"OOIP Sensitivity: {param1} vs {param2}"
    )
    
    fig_recoverable_heatmap = px.imshow(
        recoverable_matrix,
        x=[f"{p:.1f}" for p in param2_range],
        y=[f"{p:.1f}" for p in param1_range],
        labels=dict(x=f"{param2}", y=f"{param1}", color="Recoverable (bbl)"),
        color_continuous_scale="Plasma",
        title=f"Recoverable Reserves Sensitivity: {param1} vs {param2}"
    )
    
    col1, col2 = st.columns(2)
    with col1:
        st.plotly_chart(fig_ooip_heatmap, use_container_width=True)
    with col2:
        st.plotly_chart(fig_recoverable_heatmap, use_container_width=True)

# Export functionality
st.header("Export Results")

# Create a button to export results
if st.button("Generate Comprehensive Excel Report"):
    # Create a DataFrame for export
    export_data = {
        'Parameter': [
            'Reservoir Area (acres)',
            'Net Pay Thickness (ft)',
            'Porosity (%)',
            'Water Saturation (%)',
            'Oil FVF (rb/stb)',
            'Recovery Factor (%)',
            'OOIP (bbl)',
            'Recoverable Reserves (bbl)'
        ],
        'Value': [
            area,
            thickness,
            porosity,
            water_saturation,
            oil_fvf,
            recovery_factor,
            f"{ooip:,.0f}",
            f"{recoverable_reserves:,.0f}"
        ]
    }
    
    export_df = pd.DataFrame(export_data)
    
    # Create an Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Track used sheet names
        used_sheet_names = []
        
        # Base Case sheet
        sheet_name = get_unique_sheet_name("Base Case", used_sheet_names)
        export_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format the Base Case sheet
        worksheet = writer.sheets[sheet_name]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, len(export_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add sensitivity data
        if 'sensitivity_df' in locals():
            sheet_name = get_unique_sheet_name("Single Parameter Sensitivity", used_sheet_names)
            sensitivity_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the sensitivity sheet
            worksheet = writer.sheets[sheet_name]
            for col in range(1, len(sensitivity_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
        
        # Add tornado plot data
        tornado_data = {
            'Parameter': list(param_low_values.keys()),
            'OOIP Low': list(ooip_low.values()),
            'OOIP High': list(ooip_high.values()),
            'Recoverable Low': list(recoverable_low.values()),
            'Recoverable High': list(recoverable_high.values())
        }
        tornado_df = pd.DataFrame(tornado_data)
        sheet_name = get_unique_sheet_name("Tornado Plot Data", used_sheet_names)
        tornado_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format the tornado sheet
        worksheet = writer.sheets[sheet_name]
        for col in range(1, len(tornado_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add matrix data if available
        if 'ooip_matrix' in locals():
            matrix_df = pd.DataFrame(ooip_matrix, 
                                     index=[f"{p:.1f}" for p in param1_range],
                                     columns=[f"{p:.1f}" for p in param2_range])
            sheet_name = get_unique_sheet_name(f"OOIP Matrix - {param1} vs {param2}", used_sheet_names)
            matrix_df.to_excel(writer, sheet_name=sheet_name)
            
            recoverable_matrix_df = pd.DataFrame(recoverable_matrix, 
                                                index=[f"{p:.1f}" for p in param1_range],
                                                columns=[f"{p:.1f}" for p in param2_range])
            sheet_name = get_unique_sheet_name(f"Recoverable Matrix - {param1} vs {param2}", used_sheet_names)
            recoverable_matrix_df.to_excel(writer, sheet_name=sheet_name)
    
    # Create download button
    st.download_button(
        label="Download Excel Report",
        data=output.getvalue(),
        file_name="comprehensive_reserves_analysis.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.success("Comprehensive Excel report generated successfully!")

# Add footer
st.markdown("---")
st.markdown("Created by Youssif, Petroleum Reservoir Engineer")